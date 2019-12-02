# coding=utf-8
import openpyxl, xlsxwriter, os, xlrd
from win32com.client import Dispatch
from datetime import datetime
import time
import re
import shutil
import random
from sys import argv
import collections
import calendar

# 使用win32打开excel文件，才能读取公式中的数值
def just_open(filename):
    xlApp = Dispatch("ket.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()


# 获取同事评分平均分
def get_workmate_score(file_dict, start_row=16, col=17, num=3, avg_num=2):
    workmate_score = {}
    workmate_times = {}
    times = 0
    for root, dirs, files in os.walk(file_dict):
        for file in files:
            file = os.path.join(root, file)
            wb = openpyxl.load_workbook(file)
            ws = wb.worksheets[0]
            name = ws.cell(2, 2).value  # 获取员工名字
            if name not in workmate_score.keys():
                workmate_score[name] = []  # 创建同事评分列表
                workmate_times[name] = 0
                for i in range(num):
                    workmate_score[name].append(
                        ws.cell(start_row + i, col).value)
                workmate_times[name] = workmate_times[name] + 1
            else:
                workmate_times[name] = workmate_times[name] + 1
                if workmate_times[name] > 2:
                    print('%s 有超出两人人评分，请猪猪核查！' % name)
                    continue
                for i in range(num):
                    workmate_score[name][i] = (
                                                      workmate_score[name][i] + ws.cell(start_row + i,
                                                                                        col).value) / avg_num
    return workmate_score


# 获取部门评分
def get_depart_score(file_dict, start_row=8, col=17, num=6):
    depart_score = {}
    depart_times = {}
    for root, dirs, files in os.walk(file_dict):
        for file in files:
            file = os.path.join(root, file)
            wb = openpyxl.load_workbook(file)
            ws = wb.worksheets[0]
            name = ws.cell(2, 2).value  # 获取员工名字
            if name not in depart_score.keys():
                depart_score[name] = []  # 创建同事评分列表
                depart_times[name] = 0
                for i in range(num):
                    depart_score[name].append(
                        ws.cell(start_row + i, col).value)
                depart_times[name] = depart_times[name] + 1
    return depart_score


# 生成员工汇总表
def generate_person_table(name_list_table, src_table, dest_dir, table_pp, month):
    wb = openpyxl.load_workbook(name_list_table)
    ws = wb.worksheets[0]
    i = 0
    name = []
    depart = {}
    while i <= 200:
        if not re.search('备注', str(ws.cell(5 + i, 1).value)):
            # if ws.cell(5+i,18).value=='是':
            mate = ws.cell(5 + i, 2).value
            name.append(mate)
            depart[mate] = ws.cell(5 + i, 3).value
        else:
            break
        i += 1
    for n in name:
        dest_table = '%s\\%s%s.xlsx' % (dest_dir, n, table_pp)  # 写入姓名，部门，及时间
        if not os.path.exists(dest_table):
            shutil.copy2(src_table, dest_table)
            wb = openpyxl.load_workbook(dest_table)
            ws = wb.worksheets[0]
            ws.cell(2, 2, n)
            ws.cell(2, 5, depart.get(n))
            ws.cell(2, 13, '%s年%s月' % (time.strftime('%Y'), str(month)))
            wb.save(dest_table)


# 生成同事部分评分表---模拟用的
def mate_depart(name_list_table, src_table, departDict, workMateDict, month):
    generate_person_table(name_list_table, src_table, departDict, '月度部门评分', month)  # 生成部分评分
    generate_person_table(name_list_table, src_table, workMateDict, '月度同事评分-1', month)  # 生成同事1评分
    generate_person_table(name_list_table, src_table, workMateDict, '月度同事评分-2', month)  # 生成同事2评分
    choice3_0 = [3, 2, 1, 0]
    choice25_0 = [2, 1.5, 1, 0]
    for root, dirs, files in os.walk(departDict):  # 随机生成部门评分
        for file in files:
            fn = os.path.join(root, file)
            wb = openpyxl.load_workbook(fn)
            ws = wb.worksheets[0]
            for i in range(8, 11):
                ws.cell(i, 17, random.choice(choice3_0))
            for i in range(11, 14):
                ws.cell(i, 17, random.choice(choice25_0))
            wb.save(fn)
    for root, dirs, files in os.walk(workMateDict):  # 随机生成同事评分
        for file in files:
            fn = os.path.join(root, file)
            wb = openpyxl.load_workbook(fn)
            ws = wb.worksheets[0]
            for i in (16, 18):
                ws.cell(i, 17, random.choice(choice25_0))
            for i in (17,):
                ws.cell(i, 17, random.choice(choice3_0))
            wb.save(fn)


# 汇总表统计---绩效分数统计
def performance(performance_table, total_dir, month):
    wb = openpyxl.load_workbook(performance_table, data_only=True)
    ws = wb.worksheets[0]
    i = 0
    name = ''
    performance_score = {}  # {'员工':'当月分数' }
    # 获取绩效分数
    for n in range(1, 20):
        if re.findall('^%s' % str(month), ws.cell(1, n).value):
            performance_col = n  # 获取当月绩效所在列
            break
    for n in range(2, 100):

        name = ws.cell(n, 2).value
        if name is not None:
            score = ws.cell(n, performance_col).value
            if score:
                if int(score) >= 100:
                    score = 58
                else:
                    score = int(score * 0.58)
                performance_score[name] = score
        else:
            break

    for root, dirs, files in os.walk(total_dir):
        for file in files:
            fn = os.path.join(root, file)
            wb = openpyxl.load_workbook(fn)
            ws = wb.worksheets[0]
            n = ws.cell(2, 2).value
            ws.cell(4, 17, performance_score.get(n))  # 写入绩效
            wb.save(fn)


# 汇总表统计---考勤及附加分统计
def check_on_work(name_list_table, total_dir, extra_score_table, month):
    # 打开考勤汇总表
    wb = openpyxl.load_workbook(name_list_table, data_only=True)
    ws = wb.worksheets[0]
    i = 0
    name = []
    check_work = {}  # {'员工':[考勤分数，附加分加班分数，运动分，外训分，建议分]}
    while i <= 200:
        if not re.search('备注', str(ws.cell(5 + i, 1).value)):
            # if ws.cell(5 + i, 18).value == '是':
            mate = ws.cell(5 + i, 2).value
            name.append(mate)
            score = ws.cell(5 + i, 11).value
            day_off = ws.cell(5 + i, 8).value
            timeout = ws.cell(5 + i, 16).value
            if day_off:
                day_off = int(day_off)
            else:
                day_off = 0
            if timeout:
                timeout = int(timeout)
            else:
                timeout = 0
            if score:
                cw_score = 20 - day_off * 2 - timeout * 2  # 计算考勤分数，总分20
            else:
                cw_score = 0
            if cw_score < 0:
                cw_score = 0

            ot_times = ws.cell(5 + i, 12).value  # 获取加班次数
            if ot_times:
                ot_times = int(ot_times)
                if ot_times >= 8:
                    ot_score = 4
                elif ot_times >= 6:
                    ot_score = 3
                elif ot_times >= 4:
                    ot_score = 2
                else:
                    ot_score = 0
            else:
                ot_score = 0
            check_work[mate] = [int(cw_score), int(ot_score),0,0,0]
        else:
            break
        i += 1

    # 打开附加分事项统计表
    wb = xlrd.open_workbook(extra_score_table)
    ws_badminton = wb.sheet_by_name('羽毛球')  # 羽毛球名单
    ws_sujection = wb.sheet_by_name('建议')  # 建议名单
    ws_training = wb.sheet_by_name('外训')  # 培训名单

    ###羽毛球次数 sport_counter
    row = 1
    badminton_list = []
    sport_counter = collections.Counter(badminton_list)
    badminton_date_col = ws_badminton.col(1)
    for date in badminton_date_col[1:]:
        if xlrd.xldate_as_tuple(date.value, 0)[1] == int(month):
            sport_list = ws_badminton.row(row)[2:]
            for sport_name in sport_list:
                if sport_name.value:
                    badminton_list.append(sport_name.value)
                sport_counter.update(badminton_list)
                badminton_list = []
        row += 1
    ###建议次数 sujection_counter

    row = 1
    sujection_list = []
    sujection_counter = collections.Counter(sujection_list)
    sujection_month_col = ws_sujection.col(0)
    for m in sujection_month_col[1:]:
        if re.search('^%s月' % (str(month)), m.value):
            sujection_list.append(ws_sujection.row(row)[1].value)
        sujection_counter.update(sujection_list)
        row += 1

    ###外训次数 training_counter
    row = 1
    training_list = []
    training_counter = collections.Counter(training_list)
    training_month_col = ws_sujection.col(0)
    for m in training_month_col[1:]:
        if re.search('^%s月' % (str(month)), m.value):
            training_list.append(ws_training.row(row)[1].value)
        training_counter.update(training_list)
        row += 1

    # 计算运动、外训、建议分
    for n in name:
        sport_score = 0
        training_score = 0
        sujection_score = 0
        if sport_counter.get(n):
            sport_score = 0.5 * sport_counter.get(n)
            if sport_score > 2:
                sport_score = 2
            check_work[n][2] = sport_score
        if training_counter.get(n):
            training_score = 2 * training_counter.get(n)
            if training_score > 2:
                training_score = 2
            check_work[n][3] = training_score
        if sujection_counter.get(n):
            sujection_score = 5 * sujection_counter.get(n)
            check_work[n][4] = sujection_score

    for root, dirs, files in os.walk(total_dir):
        for file in files:
            fn = os.path.join(root, file)
            wb = openpyxl.load_workbook(fn)
            ws = wb.worksheets[0]
            n = ws.cell(2, 2).value
            ws.cell(5, 17, check_work.get(n)[0])  # 写入考勤分数
            ws.cell(21, 12, check_work.get(n)[1])  # 写入加班附加分
            ws.cell(22, 12, check_work.get(n)[3])  # 写入培训分
            ws.cell(23, 12, check_work.get(n)[2])  # 写入运动分
            ws.cell(25, 12, check_work.get(n)[4])  # 写入建议分
            wb.save(fn)

# 汇总同事及部门分到员工个人月度表
def total_person(
        workmate_score_dict,
        depart_score_dict,
        output_person_dict,
        workmate_start_row=16,
        depart_start_row=8,
        col=17,
        workmate_num=3,
        depart_num=6):
    for root, dirs, files in os.walk(output_person_dict):
        for file in files:
            file = os.path.join(root, file)
            wb = openpyxl.load_workbook(file)
            ws = wb.worksheets[0]
            name = ws.cell(2, 2).value
            # 写入同事平均分
            for i in range(workmate_num):
                ws.cell(workmate_start_row + i, col,
                        workmate_score_dict[name][i])
            for i in range(depart_num):
                ws.cell(depart_start_row + i, col, depart_score_dict[name][i])
            wb.save(file)




def mate_justice(justice_table,output_justice_table,month):
    wb=xlrd.open_workbook(justice_table)
    ws_mate_justice=wb.sheet_by_index(0)
    month_list=ws_mate_justice.row(0)
    for m in month_list:
        if m.value == '%s月' % str(month):
            current_month_col = month_list.index(m)
            break
    current_month_justice_name_list_1=ws_mate_justice.col(current_month_col)[2:]
    current_month_justice_name_list_2 = ws_mate_justice.col(current_month_col+1)[2:]
    justiced_list=ws_mate_justice.col(2)[2:]
    justice={}
    for name in current_month_justice_name_list_1:
        index=current_month_justice_name_list_1.index(name)
        name = name.value
        if name and name != '/':
            if name not in justice.keys():
                justice[name]=justiced_list[index].value
            else:
                justice[name]=justice[name] + '、' + justiced_list[index].value
    for name in current_month_justice_name_list_2:
        index = current_month_justice_name_list_2.index(name)
        name = name.value
        if name and name!='/':
            if name not in justice.keys():
                justice[name] = justiced_list[index].value
            else:
                justice[name] = justice[name] + '、' + justiced_list[index].value
    title1='%s月月度互评名单' % str(month)
    wbook = xlsxwriter.Workbook(output_justice_table)  # 打开工作簿
    wsheet = wbook.add_worksheet('%s月月度互评名单' % str(month))
    bold = wbook.add_format({
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        # 'fg_color': '#FCD5B4',  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })
    wsheet.set_column('B:C',25)
    wsheet.merge_range('A1:C1', title1, bold)
    wsheet.write('A2','评价人',bold)
    wsheet.merge_range('B2:C2', '被评价人', bold)
    row = 2
    for name in justice.keys():
        row += 1
        wsheet.write('A%d' % (row),name,bold)
        #print(name)
        wsheet.merge_range('B%d:C%d' % (row,row), justice[name], bold)
    wbook.close()

# 生成总表
def total_everyone(person_dict, jixiao_table,output_everyone_file,ex_month_total_table,month):
    first_row = ['序号', '姓名', '绩效/考勤', '工作/综合(部门)', '综合(同事)', '附加分', '总分','绩效明细','进步分数']
    number = 0
    score = {}
    # 读取绩效
    jixiao={}
    wb_jixiao=xlrd.open_workbook(jixiao_table)
    ws_jixiao=wb_jixiao.sheet_by_name('绩效统计')
    month_col=ws_jixiao.row(0)
    for m in month_col:
        index=month_col.index(m)
        m=m.value
        if m:
            if re.search('%s月' % (str(month)),m):
                month_index=index
                break

    name_jixiao=ws_jixiao.col(1)
    jixiao_score=ws_jixiao.col(month_index)
    for i,n in enumerate(name_jixiao):
        n=n.value
        jixiao[n]=jixiao_score[i].value

    # 读取上月分数
    wb_ex = xlrd.open_workbook(ex_month_total_table)
    ws_ex = wb_ex.sheet_by_index(0)
    name_ex = ws_ex.col(1)[1:]
    total_score_ex = ws_ex.col(6)[1:]
    ex_score = {}
    for i, n in enumerate(name_ex):
        n = n.value
        if n and n != '姓名':
            ex_score[n] = total_score_ex[i].value

    # 读取加权合计以及总分
    for root, dirs, files in os.walk(person_dict):
        for file in files:
            file = os.path.join(root, file)
            just_open(file)
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.worksheets[0]
            number += 1
            name = ws.cell(2, 2).value
            score1 = ws.cell(6, 3).value
            score2 = ws.cell(14, 3).value
            score3 = ws.cell(19, 3).value
            score4 = ws.cell(26, 3).value
            score5 = ws.cell(27, 3).value
            score6 = jixiao.get(name)
            if ex_score.get(name):
                score7 = score5 - ex_score.get(name)
            else:
                score7=None
            # print(score5)
            score[name] = [number, name, score1, score2, score3, score4, score5,score6,score7]

    wbook = xlsxwriter.Workbook(output_everyone_file)  # 打开工作簿
    wsheet = wbook.add_worksheet('月度员工评分汇总')
    bold = wbook.add_format({
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        # 'fg_color': '#F4B084',  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })
    wsheet.write_row('A1', first_row, bold)
    row = 1
    for name in score:
        row += 1
        wsheet.write_row('A%d' % (row), score[name], bold)
    wbook.close()


if __name__ == '__main__':
    time_now = datetime.now()
    month = input('请输入需要统计的月份：')
    if not month:
        month = time_now.month

    if int(month) < 10:
        mon0 = str(month)
        mon0 = '0' + mon0
    else:
        mon0=str(month)
    ex_month=int(month)-1
    if ex_month==0:
        ex_month=12
    else:
        ex_month=ex_month

    root = os.getcwd()
    workMateDict = r'%s\%s月\同事评分表' % (root, month)
    departDict = r'%s\%s月\部门评分表' % (root, month)
    personDict = r'%s\%s月\汇总表' % (root, month)
    ex_everyonePath = r'%s\%s月\%s月月度员工总分表.xlsx' % (root, month, ex_month)
    everyonePath = r'%s\%s月\%s月月度员工总分表.xlsx' % (root, month, month)
    name_list_table = r'%s\%s月\考勤汇总2019.%s-兑正.xlsx' % (root, month, mon0)
    performance_table = r'%s\%s月\绩效统计.xlsx' % (root, month)
    src_table = r'%s\月度员工评选.xlsx' % (root)
    extra_table = r'%s\%s月\附加分项.xlsx' % (root, month)
    justice_table = r'%s\%s月\互评名单.xlsx' % (root, month)
    output_justice_table = r'%s\%s月\%s月互评名单-统计.xlsx' % (root, month,month)

    # 不加任何选项，只生成个人汇总表
    generate_person_table(name_list_table, src_table, personDict, '月度汇总', month)  # 生成汇总表

    if argv[1].startswith('-'):
        options = argv[1][1:]
        if 'm' in options:  # 模拟生成同事部门评分表模式
            mate_depart(name_list_table, src_table, departDict, workMateDict, month)  # 生成部门同事评分
        if 'a' in options or 't' in options:
            workmateScore = get_workmate_score(workMateDict, 16, 17, 3, 2)  # 获取同事评分
            departScore = get_depart_score(departDict, 8, 17)  # 获取部门评分
            total_person(workmateScore, departScore, personDict, 16, 8, 17, 3, 6)  # 汇总同事及部门分到个人汇总表
        if 'a' in options or 'c' in options:
            check_on_work(name_list_table, personDict, extra_table, month)  # 汇总考勤及加班分到个人汇总表
        if 'a' in options or 'p' in options:
            performance(performance_table, personDict, month)
        if 'a' in options or 'e' in options:
            total_everyone(personDict, performance_table, everyonePath, ex_everyonePath,month)
        if 'a' in options or 'j' in options:
            mate_justice(justice_table,output_justice_table,month)  #生成月度互评名单
        for op in options:  # 判断错误选项
            if op not in 'acetmpj':
                print('invalid options %s' % (op,))
