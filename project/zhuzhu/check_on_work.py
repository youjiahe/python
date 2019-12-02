# coding=utf-8
import openpyxl, xlsxwriter, os, xlrd
from win32com.client import Dispatch
from datetime import datetime
import time
import re
import shutil
import random
from sys import argv
import collections
import calendar
import requests
import xlwings as xw

# 使用win32打开excel文件，才能读取公式中的数值
def just_open(filename):
    xlApp = Dispatch("ket.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()

def copy_sheet(wb1_path,wsheet1,wb2_path,wsheet2):
    wb1 = xw.Book(wb1_path)
    wb2 = xw.Book(wb2_path)

    ws1 = wb1.sheets(int(wsheet1))

    ws1.api.Copy(Before=wb2.sheets(wsheet2).api)
    wb2.save()
    wb2.app.quit()

def generate_check_on_work_table(name_list_table,output_check_on_work_table,year,month):
    #获取数据
    year=int(year)
    month=int(month)
    wb = openpyxl.load_workbook(name_list_table, data_only=True)
    ws = wb.worksheets[0]
    i = 0
    name = []
    while i <= 200:
        if not re.search('备注', str(ws.cell(5 + i, 1).value)):
            mate = ws.cell(5 + i, 2).value
            name.append(mate)
        else:
            break
        i+=1
    a=''
    title_list = ['Name', '', '']
    empty_list = [a for i in range(calendar.monthlen(year, month))]
    for i in range(1, calendar.monthlen(year,month) + 1):
        day = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
        title_list.append('%d %s' % (i, day.get(calendar.weekday(year, month, i))))

    #开始生成表格--打开工作簿
    wbook = xlsxwriter.Workbook(output_check_on_work_table)  # 打开工作簿
    wsheet = wbook.add_worksheet('明细')
    border_color='#0000FD'
    fg_color1 = '#FFFFFF'
    fg_color2 = '#FCD5B4'
    fg_color=fg_color1
    ###设置样式
    bold_title = wbook.add_format({
        'font_size':14,
        'font_color':'#1205BB',
        'bold': True,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'border_color':border_color,
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        # 'fg_color': '#FCD5B4',  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })

    bold_name1 = wbook.add_format({
        'font_size': 14,
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'border_color': border_color,
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        'fg_color': fg_color1, # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })
    bold_name2 = wbook.add_format({
        'font_size': 14,
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'border_color': border_color,
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        'fg_color': fg_color2, # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })
    bold_time1 = wbook.add_format({
        'font_size': 13,
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'border_color': border_color,
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        'fg_color': fg_color1,  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })
    bold_time2 = wbook.add_format({
        'font_size': 13,
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'border_color': border_color,
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        'fg_color': fg_color2,  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })

    bold_ban1 = wbook.add_format({
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'border_color': border_color,
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        'fg_color': fg_color1,  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })

    bold_ban2 = wbook.add_format({
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'border_color': border_color,
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        'fg_color': fg_color2,  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })

    bold_empty1 = wbook.add_format({
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'border_color': border_color,
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        'fg_color': fg_color1,  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })
    bold_empty2 = wbook.add_format({
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        'border_color': border_color,
        'align': 'center',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        'fg_color': fg_color2,  # 单元格背景颜色
        'text_wrap': True,  # 是否自动换行
    })

    bold_name=bold_name1
    bold_time=bold_time1
    bold_ban=bold_ban1
    bold_empty=bold_empty1

    #设置行高列宽
    wsheet.set_row(0, 28)  #title行
    for i in range(len(name)*4):
        wsheet.set_row(i+1, 17)
    wsheet.set_column(0,13)
    ###开始写表

    wsheet.write_row('A1',title_list,bold_title)   #写title
    row=2
    for i in range(len(name)*4):
        # 交换颜色
        if i % 4 == 0 and i > 0:
            if bold_empty == bold_empty1:
                bold_empty = bold_empty2
            else:
                bold_empty = bold_empty1
        wsheet.write_row('D%d' % (row), empty_list, bold_empty)  #写空内容，就是打卡时间单元格
        row+=1

    row=2
    for n in name:
        wsheet.merge_range('A%d:A%d' % (row,row+3),n,bold_name)              #写入名字
        wsheet.merge_range('B%d:B%d' % (row,row+1),'上午',bold_time)          #写入上下午
        wsheet.merge_range('B%d:B%d' % (row+2, row + 3), '下午', bold_time)   #写入上下午
        wsheet.write('C%d' % (row), '上班', bold_ban)                         #写入上下班
        wsheet.write('C%d' % (row + 1), '下班', bold_ban)
        wsheet.write('C%d' % (row + 2), '上班', bold_ban)
        wsheet.write('C%d' % (row + 3), '下班', bold_ban)
        row+=4
        #交换颜色
        if bold_ban == bold_ban1:
            bold_ban = bold_ban2
        else:
            bold_ban = bold_ban1
        if bold_name == bold_name1:
            bold_name = bold_name2
        else:
            bold_name = bold_name1
        if bold_time==bold_time1:
            bold_time = bold_time2
        else:
            bold_time = bold_time1
    wbook.close()  #保存
    copy_sheet(output_check_on_work_table,1,name_list_table,2)
    os.remove(output_check_on_work_table)
    time.sleep(3)
    load_data_into_check_work_detal_table(src_check_on_work_table, name_list_table)

def check_on_work_total(check_on_work_detal_table,output_detal_tabale,year,month,src_check_on_work_table):
    def day_off_justice(col,wbook,wsheet):
        r=0
        counter=0
        for c in wsheet.col(col)[1:]:
            if c.value:
                # print(col)
                counter+=1
                if counter>=7:
                    r=1
                    break
        return r
    wbook=xlrd.open_workbook(check_on_work_detal_table)
    wsheet=wbook.sheet_by_name('明细')
    name=[]
    name_list=wsheet.col(0)
    for n in name_list:
        if n.value and n.value!='Name':
          name.append(n.value)
    title_list=wsheet.row(0)
    check_on_work_dict={}  #{'name':[迟到次数，早退次数，缺打卡次数]}
    comment={}   #{'name':{'迟到':'9-12,9-30','早退':'9-1,9-2','缺打卡':'9-12'},'name1':{'迟到':'','早退':'','缺打卡':''}}
    row=1

    # for i in range(1, calendar.monthlen(year, month) + 1):
    #     if calendar.weekday(year,month,i)==5:

    weekday_on_duty_time=datetime.strptime('09:00','%H:%M')
    weekday_off_duty_time = datetime.strptime('18:00', '%H:%M')
    Sat_on_duty_time = datetime.strptime('09:30', '%H:%M')
    Sat_off_duty_time = datetime.strptime('12:00', '%H:%M')
    while row<len(name_list):
        on_duty_time_list = wsheet.row(row)
        Sat_off_duty_time_list = wsheet.row(row + 1)
        off_duty_time_list = wsheet.row(row+3)
        check_on_work_dict[name_list[row].value] = [0,0,0]
        comment[name_list[row].value] = {'迟到':'','早退':'','缺卡':''}
        # 迟到次数统计
        for on_duty_time in on_duty_time_list:
            index = on_duty_time_list.index(on_duty_time)
            da=re.search('[0-9]{1,2}', title_list[index].value)
            if da:
                date_index = str(month) + '-' + da.group() + ','  # 生成日期
            # print(on_duty_time.value)
            if on_duty_time.value:
                if re.search('[0-9]{1,2}:[0-9]{1,2}',str(on_duty_time.value)):
                    on_duty_time=datetime.strptime(on_duty_time.value,'%H:%M')
                    if re.findall('(一|二|三|四|五|日)', title_list[index].value):
                        if on_duty_time>weekday_on_duty_time :  #工作日
                            check_on_work_dict[name_list[row].value][0]+=1    # 迟到次数统计
                            comment[name_list[row].value]['迟到']+=date_index
                    elif re.findall('(六)',title_list[index].value):  #周六
                        if on_duty_time>Sat_on_duty_time:
                            check_on_work_dict[name_list[row].value][0] += 1  # 迟到次数统计
                            comment[name_list[row].value]['迟到'] += date_index
                elif on_duty_time.ctype==3:
                    on_duty_time=xlrd.xldate_as_tuple(on_duty_time.value,0)
                    on_duty_time=datetime.strptime('%d:%d' % (on_duty_time[3],on_duty_time[4]),'%H:%M')
                    if re.findall('(一|二|三|四|五|日)', title_list[index].value):
                        if on_duty_time>weekday_on_duty_time :  #工作日
                            check_on_work_dict[name_list[row].value][0]+=1    # 迟到次数统计
                            comment[name_list[row].value]['迟到']+=date_index
                    elif re.findall('(六)',title_list[index].value):  #周六
                        if on_duty_time>Sat_on_duty_time:
                            check_on_work_dict[name_list[row].value][0] += 1  # 迟到次数统计
                            comment[name_list[row].value]['迟到'] += date_index
                            
            elif on_duty_time.value =='' and index>2:
                if day_off_justice(index, wbook, wsheet) == 0:
                    continue
                else:
                    check_on_work_dict[name_list[row].value][2] += 1  # 缺卡次数t
                    comment[name_list[row].value]['缺卡'] += date_index
            
            
        # 早退次数统计
        for off_duty_time in off_duty_time_list:
            index = off_duty_time_list.index(off_duty_time)
            da = re.search('[0-9]{1,2}', title_list[index].value)
            if da:
                date_index = str(month) + '-' + da.group() + ','  # 生成日期
            if off_duty_time.value:
                if re.search('[0-9]{1,2}:[0-9]{1,2}',str(off_duty_time.value)):
                    off_duty_time=datetime.strptime(off_duty_time.value,'%H:%M')
                    if re.findall('(一|二|三|四|五|日)', title_list[index].value):
                        if off_duty_time<weekday_off_duty_time:
                            check_on_work_dict[name_list[row].value][1] += 1  # 早退次数统计
                            comment[name_list[row].value]['早退'] += date_index
                elif off_duty_time.ctype==3:
                    off_duty_time=xlrd.xldate_as_tuple(off_duty_time.value,0)
                    off_duty_time=datetime.strptime('%d:%d' % (off_duty_time[3],off_duty_time[4]),'%H:%M')
                    if re.findall('(一|二|三|四|五|日)', title_list[index].value):
                        if off_duty_time<weekday_off_duty_time:
                            check_on_work_dict[name_list[row].value][1] += 1  # 早退次数统计
                            comment[name_list[row].value]['早退'] += date_index
                            
            elif off_duty_time.value =='' and index>2:
                if re.findall('(六)', title_list[index].value):
                    continue
                if day_off_justice(index, wbook, wsheet) == 0:
                    continue
                else:
                    check_on_work_dict[name_list[row].value][2] += 1  # 缺卡次数t
                    comment[name_list[row].value]['缺卡'] += date_index

        # 周六早退统计
        for sat_off_duty_time in Sat_off_duty_time_list:
            index = Sat_off_duty_time_list.index(sat_off_duty_time)
            da = re.search('[0-9]{1,2}', title_list[index].value)
            if da:
                date_index = str(month) + '-' + da.group() + ','  # 生成日期
            if sat_off_duty_time.value:
                if re.search('[0-9]{1,2}:[0-9]{1,2}',str(sat_off_duty_time.value)):
                    sat_off_duty_time=datetime.strptime(sat_off_duty_time.value,'%H:%M')
                    if re.findall('六',title_list[index].value):
                        if sat_off_duty_time<Sat_off_duty_time:
                            check_on_work_dict[name_list[row].value][1] += 1  # 早退次数统计
                            comment[name_list[row].value]['早退'] += date_index
                            
                elif sat_off_duty_time.ctype == 3:
                    sat_off_duty_time=xlrd.xldate_as_tuple(sat_off_duty_time.value,0)
                    sat_off_duty_time=datetime.strptime('%d:%d' % (sat_off_duty_time[3],sat_off_duty_time[4]),'%H:%M')
                    if re.findall('六',title_list[index].value):
                        if sat_off_duty_time<Sat_off_duty_time:
                            check_on_work_dict[name_list[row].value][1] += 1  # 早退次数统计
                            comment[name_list[row].value]['早退'] += date_index
                            
                           
            elif sat_off_duty_time.value =='' and index>2:
                if re.findall('(一|二|三|四|五|日)', title_list[index].value):
                    continue
                if day_off_justice(index, wbook, wsheet) == 0:
                    continue
                else:
                    check_on_work_dict[name_list[row].value][2] += 1  # 缺卡次数t
                    comment[name_list[row].value]['缺卡'] += date_index
        row+=4
    # print(check_on_work_dict)
    # for co in comment:
    #     print('%s : %s ' % (co,comment[co]))

    wb_check_on_work_detal=xlsxwriter.Workbook(output_detal_tabale)
    ws_check_on_work_detal=wb_check_on_work_detal.add_worksheet('%s月考勤次数汇总' % (month))
    bold = wb_check_on_work_detal.add_format({
        'bold': False,  # 字体加粗
        'border': 1,  # 单元格边框宽度
        #'border_color': border_color,
        'align': 'left',  # 水平对齐方式
        'valign': 'vcenter',  # 垂直对齐方式
        #'fg_color': fg_color1,  # 单元格背景颜色
        'text_wrap': False,  # 是否自动换行
    })

    title_detal=['姓名','迟到次数','早退次数','缺卡次数','备注']
    ws_check_on_work_detal.write_row('A1', title_detal, bold)
    comments={}
    row=2
    for n in name:
        comments[n]='迟到：%s; 早退：%s； 缺卡：%s '  % (comment[n]['迟到'],comment[n]['早退'],comment[n]['缺卡'])
        # comments[n] = '迟到：%s; 早退：%s；' % (comment[n]['迟到'], comment[n]['早退'])

    for n in name:
        if n:
            print(n)
            write_list=[n,
                        check_on_work_dict[n][0],
                        check_on_work_dict[n][1],
                        check_on_work_dict[n][2],
                        comments[n]]
            print(write_list)
            ws_check_on_work_detal.write_row('A%d' % row,write_list,bold)
        row+=1
    ws_check_on_work_detal.set_column(1, 20)
    ws_check_on_work_detal.set_column(4, 50)
    wb_check_on_work_detal.close()

def load_data_into_check_work_detal_table(src_check_on_work_table,name_list_table):
    def date_tuple(data):
        if data:
            tup = xlrd.xldate_as_tuple(data,0)
            tup = '{}:{:0>2d}'.format(tup[3],tup[4])
        else:
            tup = ''
        return tup
    wbooks=xlrd.open_workbook(src_check_on_work_table)
    work_time={}
    #{'姓名':{'1':['9:00','12:00','13:30','18:00'],'2':['9:00','12:00','13:30','18:00'],'3':['9:00','12:00','13:30','18:00'],}}
    for wb in wbooks.sheets():
        wb=wb.name
        if wb != '排班记录表' and wb != '考勤汇总表':
            wsheet = wbooks.sheet_by_name('%s' % (wb))
            name_row = wsheet.row(3)
            index_morning_on=1
            index_morning_off=3
            index_aftering_on=6
            index_aftering_off=8
            for name in name_row:
                name_index = name_row.index(name)
                if name_index in (9,24,39):
                    name = name.value.strip()
                    work_time[name]={} #初始化个人考勤数据字典
                    wsheet_cols=wsheet.col(0)  #读取考勤原始表有多少行数据
                    row=12
                    date = 1
                    work_time[name][date] = []
                    for col in wsheet_cols[12:]:   #从第12行（0起始）开始读取考勤数据
                        person_check_on_work_src_data=wsheet.row(row)
                        morning_on=date_tuple(person_check_on_work_src_data[index_morning_on].value)
                        morning_off=date_tuple(person_check_on_work_src_data[index_morning_off].value)
                        aftering_on=date_tuple(person_check_on_work_src_data[index_aftering_on].value)
                        aftering_off=date_tuple(person_check_on_work_src_data[index_aftering_off].value)
                        if aftering_off=='':
                            index_aftering_off_again = index_aftering_off + 2
                            aftering_off = date_tuple(person_check_on_work_src_data[index_aftering_off_again].value)
                            if aftering_off=='':
                                index_aftering_off_again_again = index_aftering_off_again + 2
                                aftering_off = date_tuple(person_check_on_work_src_data[index_aftering_off_again_again].value)
                        work_time[name][date]=[morning_on,morning_off,aftering_on,aftering_off]
                        date += 1
                        row += 1
                    index_morning_on += 15
                    index_morning_off += 15
                    index_aftering_on += 15
                    index_aftering_off += 15

    # 写数据导考勤明细表
    ## 读取姓名名单
    wb_dest_table_src=xlrd.open_workbook(name_list_table)
    ws_dest_table_src=wb_dest_table_src.sheet_by_name('明细')
    name_col=ws_dest_table_src.col(0)
    name_list=[]
    for name in name_col:
        name=name.value
        if name!='Name' and name:
            name_list.append(name)
    # 根据名字写入数据
    wb_dest_table=openpyxl.load_workbook(name_list_table)
    ws_dest_table=wb_dest_table.worksheets[1]
    row=2
    col=4
    for n in name_list:
        # print(n)
        # if n:
        if work_time.get(n):
            for time_dic in work_time.get(n).values():
                # print(time_dic)
                for time_str in time_dic:
                    # print(time_str)
                    ws_dest_table.cell(row,col,time_str)
                    row += 1
                col += 1
                row -= 4
        row += 4
        col = 4
    wb_dest_table.save(name_list_table)
    # return work_time


if __name__ == '__main__':
    time_now = datetime.now()
    year = input('请输入年份(YEAR)：')
    month = input('请输入月份(MONTH)：')
    if not year:
        month = time_now.year
    if not month:
        month = time_now.month

    if int(month) < 10:
        mon0 = str(month)
        mon0 = '0' + mon0
    else:
        mon0 = str(month)

    root = os.getcwd()
    name_list_table = r'%s\%s月\考勤汇总%s.%s-兑正.xlsx' % (root, month,year, mon0)
    output_check_on_work_table=r'%s\%s月\%s月考勤明细.xlsx' % (root,month,month)
    deltal_times_table=r'%s\%s月\%s月考勤次数统计.xlsx' % (root,month,month)
    src_check_on_work_table = r'%s\%s月\考勤报表.%s.xls' % (root,month,month)

    if argv[1].startswith('-'):
        options = argv[1][1:]
        if 'g' in options:
            generate_check_on_work_table(name_list_table,output_check_on_work_table,year,month)
        if 'c' in options:
            check_on_work_total(name_list_table,deltal_times_table,year,month,src_check_on_work_table)
        for op in options:  # 判断错误选项
            if op not in 'gc':
                print('invalid options %s' % (op,))
